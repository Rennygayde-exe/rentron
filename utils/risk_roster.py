from __future__ import annotations

import json
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RISK_SOURCE_FILE = Path("data/High Risk TMH.xlsx")
RISK_ROSTER_DIR = Path("riskroster")

FIELD_MAP = [
    ("name", "Name"),
    ("risk_factor", "Risk Factor"),
    ("discord_username", "Discord Username"),
    ("location", "Location"),
    ("date_of_risk", "Date of Risk"),
    ("risk_behaviors", "Risk Behaviors"),
    ("pocs", "POCs to Help"),
    ("sheet_link", "Link to Sheet"),
    ("last_contacted", "Date Last Contacted"),
]

DATE_FIELDS = {"date_of_risk", "last_contacted"}
ENTRY_FIELDS = [field for field, _ in FIELD_MAP]
_SLUG_PATTERN = re.compile(r"[^a-z0-9]+")


class RiskRosterError(Exception):
    """Base error for risk roster operations."""


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return str(value)


def _format_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return None


def _slugify(value: str | None) -> str:
    base = (value or "entry").strip().lower()
    base = _SLUG_PATTERN.sub("-", base).strip("-")
    return base or "entry"


def _row_to_entry(header: list[str], row: tuple[Any, ...]) -> dict[str, Any] | None:
    row_map = {header[idx]: cell for idx, cell in enumerate(row) if idx < len(header) and header[idx]}
    entry: dict[str, Any] = {}

    for field, column in FIELD_MAP:
        raw_value = row_map.get(column)
        if field in DATE_FIELDS:
            entry[field] = _format_date(raw_value)
        else:
            entry[field] = _clean_text(raw_value)

    primary_fields = (entry.get("name"), entry.get("discord_username"))
    if not any(primary_fields):
        return None
    return entry


def build_entries_from_sheet() -> list[dict[str, Any]]:
    if not RISK_SOURCE_FILE.exists():
        raise FileNotFoundError(f"Roster source file not found at {RISK_SOURCE_FILE}")

    workbook = load_workbook(RISK_SOURCE_FILE, data_only=True)
    worksheet = workbook.active

    header = [
        (cell.value or "").strip() if isinstance(cell.value, str) else cell.value
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
    ]

    entries: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        entry = _row_to_entry(header, row)
        if entry:
            entries.append(entry)

    slug_counts: dict[str, int] = {}
    for entry in entries:
        base = _slugify(entry.get("discord_username") or entry.get("name"))
        count = slug_counts.get(base, 0) + 1
        slug_counts[base] = count
        slug = base if count == 1 else f"{base}-{count}"
        entry["id"] = slug

    return entries


def sync_roster_files(entries: list[dict[str, Any]]) -> list[Path]:
    if not entries:
        return []

    RISK_ROSTER_DIR.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for entry in entries:
        file_path = RISK_ROSTER_DIR / f"{entry['id']}.json"
        with file_path.open("w", encoding="utf-8") as f:
            json.dump(entry, f, indent=2)
        paths.append(file_path)
    return paths


def load_saved_entries() -> list[dict[str, Any]]:
    if not RISK_ROSTER_DIR.exists():
        return []

    entries: list[dict[str, Any]] = []
    for file_path in sorted(RISK_ROSTER_DIR.glob("*.json")):
        try:
            with file_path.open("r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    entries.append(data)
        except json.JSONDecodeError:
            continue
    entries.sort(key=lambda e: (e.get("name") or e.get("discord_username") or "").lower())
    return entries


def format_entry_table(entry: dict[str, Any]) -> str:
    rows = [
        ("Name", entry.get("name") or "N/A"),
        ("Discord", entry.get("discord_username") or "N/A"),
        ("Risk Factor", entry.get("risk_factor") or "N/A"),
        ("Location", entry.get("location") or "N/A"),
        ("Risk Date", entry.get("date_of_risk") or "N/A"),
        ("Behaviors", entry.get("risk_behaviors") or "N/A"),
        ("POCs", entry.get("pocs") or "N/A"),
        ("Last Contacted", entry.get("last_contacted") or "N/A"),
        ("Sheet Link", entry.get("sheet_link") or "N/A"),
    ]
    key_width = max(len(label) for label, _ in rows)
    lines = [f"{label.ljust(key_width)} : {value}" for label, value in rows]
    return "\n".join(lines)


def refresh_roster_from_sheet() -> list[dict[str, Any]]:
    entries = build_entries_from_sheet()
    sync_roster_files(entries)
    return entries


def _clean_field(field: str, value: Any) -> Any:
    if field in DATE_FIELDS:
        return _format_date(value)
    return _clean_text(value)


def _normalize_entry(data: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for field in ENTRY_FIELDS:
        normalized[field] = _clean_field(field, data.get(field))
    return normalized


def _dedupe_slug(base: str, existing_ids: set[str]) -> str:
    if base not in existing_ids:
        return base
    idx = 2
    while True:
        slug = f"{base}-{idx}"
        if slug not in existing_ids:
            return slug
        idx += 1


def _ensure_notes(entry: dict[str, Any]) -> list[dict[str, Any]]:
    notes = entry.get("notes")
    if isinstance(notes, list):
        cleaned = [note for note in notes if isinstance(note, dict)]
    else:
        cleaned = []
    entry["notes"] = cleaned
    return cleaned


def load_entry(entry_id: str) -> dict[str, Any] | None:
    file_path = RISK_ROSTER_DIR / f"{entry_id}.json"
    if not file_path.exists():
        return None
    with file_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
        if isinstance(data, dict):
            data.setdefault("id", entry_id)
            _ensure_notes(data)
            return data
    return None


def save_entry(entry: dict[str, Any]) -> Path:
    entry_id = entry.get("id")
    if not entry_id:
        raise RiskRosterError("Entry is missing an id.")
    RISK_ROSTER_DIR.mkdir(parents=True, exist_ok=True)
    file_path = RISK_ROSTER_DIR / f"{entry_id}.json"
    with file_path.open("w", encoding="utf-8") as f:
        json.dump(entry, f, indent=2)
    return file_path


def add_entry(data: dict[str, Any]) -> dict[str, Any]:
    entry = _normalize_entry(data)
    if not entry.get("name") and not entry.get("discord_username"):
        raise RiskRosterError("At least a name or Discord username is required.")

    existing_ids = {path.stem for path in RISK_ROSTER_DIR.glob("*.json")} if RISK_ROSTER_DIR.exists() else set()
    preferred = entry.get("discord_username") or entry.get("name")
    slug = _dedupe_slug(_slugify(preferred), existing_ids)
    entry["id"] = slug
    entry.setdefault("notes", [])
    save_entry(entry)
    return entry


def update_entry(entry_id: str, updates: dict[str, Any]) -> dict[str, Any]:
    entry = load_entry(entry_id)
    if entry is None:
        raise RiskRosterError(f"No roster entry found for'{entry_id}'.")
    if not updates:
        return entry

    for field, value in updates.items():
        if field not in ENTRY_FIELDS:
            continue
        entry[field] = _clean_field(field, value)
    save_entry(entry)
    return entry


def remove_entry(entry_id: str) -> dict[str, Any]:
    entry = load_entry(entry_id)
    if entry is None:
        raise RiskRosterError(f"No roster entry found for '{entry_id}'.")
    file_path = RISK_ROSTER_DIR / f"{entry_id}.json"
    try:
        file_path.unlink()
    except FileNotFoundError:
        pass
    return entry


def _next_note_id(notes: list[dict[str, Any]]) -> int:
    max_id = 0
    for note in notes:
        try:
            max_id = max(max_id, int(note.get("id", 0)))
        except (TypeError, ValueError):
            continue
    return max_id + 1


def _timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()


def list_notes(entry_id: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    entry = load_entry(entry_id)
    if entry is None:
        raise RiskRosterError(f"No roster entry found for '{entry_id}'.")
    notes = _ensure_notes(entry)
    return entry, notes


def add_note(entry_id: str, author_name: str, author_id: int, content: str) -> tuple[dict[str, Any], dict[str, Any]]:
    if not content.strip():
        raise RiskRosterError("Note content cant be empty.")
    entry, notes = list_notes(entry_id)
    note_id = _next_note_id(notes)
    note = {
        "id": note_id,
        "author": author_name,
        "author_id": author_id,
        "content": content.strip(),
        "created_at": _timestamp(),
        "updated_at": None,
    }
    notes.append(note)
    save_entry(entry)
    return entry, note


def edit_note(entry_id: str, note_id: int, content: str) -> tuple[dict[str, Any], dict[str, Any]]:
    if not content.strip():
        raise RiskRosterError("Note content cant be empty.")
    entry, notes = list_notes(entry_id)
    for note in notes:
        if int(note.get("id", -1)) == int(note_id):
            note["content"] = content.strip()
            note["updated_at"] = _timestamp()
            save_entry(entry)
            return entry, note
    raise RiskRosterError(f"Note {note_id} not found for entry '{entry_id}'.")


def remove_note(entry_id: str, note_id: int) -> tuple[dict[str, Any], dict[str, Any]]:
    entry, notes = list_notes(entry_id)
    for idx, note in enumerate(notes):
        if int(note.get("id", -1)) == int(note_id):
            removed = notes.pop(idx)
            save_entry(entry)
            return entry, removed
    raise RiskRosterError(f"Note {note_id} not found for entry '{entry_id}'.")
